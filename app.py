import io
import re
from datetime import datetime
from pathlib import Path
from html import escape
from typing import List, Dict, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

PAS_YELLOW = "#FFD400"
PAS_BLACK = "#0A0A0A"
APP_VERSION = "v1.0 Prototype Build"
OUTPUT_COLUMNS = [
    "Site",
    "Order No",
    "Contract No",
    "Fleet / Stock No",
    "Description",
    "Qty",
    "On Hire Date",
    "Weekly Rate",
    "Net Weekly",
]

st.set_page_config(page_title="PAS Vendor PDF Converter", page_icon="pas_logo.png", layout="wide")

# =========================
# PAS UI STYLE - copied to match existing PAS apps
# =========================
st.markdown(
    f"""
    <style>
    .stApp {{ background: #f5f5f5; color: #0A0A0A; }}
    section[data-testid="stSidebar"] {{
        background: {PAS_BLACK};
        color: white;
        padding-top: 1.45rem;
    }}
    section[data-testid="stSidebar"] * {{ color: white; }}
    section[data-testid="stSidebar"] img {{
        margin-top: 0.15rem;
        border-radius: 14px;
    }}
    .block-container {{
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1500px;
    }}

    .pas-hero {{ display:flex; align-items:center; gap:16px; background: linear-gradient(100deg, #08090b 0%, #151718 70%, #c9aa00 130%) !important; border-radius: 16px !important; padding: 12px 22px !important; margin: 0 0 18px 0 !important; box-shadow: 0 9px 25px rgba(0,0,0,.13) !important; min-height:60px; }}
    .pas-hero-logo {{ width:37px; height:37px; border-radius:7px; background:{PAS_YELLOW}; color:#000; display:inline-flex; align-items:center; justify-content:center; font-weight:950; font-size:14px; letter-spacing:-1px; }}
    .pas-hero-text {{ color:#fff; font-size:18px; font-weight:950; letter-spacing:-.02em; }}
    .pas-hero-dot {{ color:#fff; opacity:.8; margin: 0 7px; }}
    .pas-hero-version {{ color:{PAS_YELLOW}; font-weight:950; }}

    .pas-upload-card {{ background:#fff; border:1px solid #e5e7eb; border-radius:18px; box-shadow:0 5px 18px rgba(15,23,42,.08); padding:16px 18px 14px; margin-bottom:14px; }}
    .pas-upload-title {{ color:#0A0A0A; font-size:16px; font-weight:950; margin-bottom:10px; }}

    div[data-testid="stFileUploader"] {{ margin:0 !important; }}
    div[data-testid="stFileUploader"] label {{ display:none !important; }}
    div[data-testid="stFileUploader"] section {{ background:#f4f6f8 !important; border:1px solid #dfe4ea !important; border-radius:11px !important; min-height:52px !important; padding:8px 10px !important; }}
    div[data-testid="stFileUploader"] section * {{ color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] button {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #d7dce3 !important; border-radius:10px !important; font-weight:900 !important; box-shadow:0 2px 8px rgba(0,0,0,.06) !important; }}
    div[data-testid="stFileUploader"] svg {{ color:#0A0A0A !important; fill:currentColor !important; stroke:currentColor !important; }}
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ display: none !important; }}
    div[data-testid="stFileUploaderDropzone"] {{ background: transparent !important; border: 0 !important; padding: 0 !important; min-height: 0 !important; }}
    div[data-testid="stFileUploaderDropzoneInstructions"] {{ display: none !important; }}
    div[data-testid="stFileUploader"] section {{ background: transparent !important; border: 0 !important; min-height: 0 !important; padding: 0 !important; }}

    .stButton > button, .stDownloadButton > button {{
        background: {PAS_YELLOW} !important;
        color: {PAS_BLACK} !important;
        border: 1px solid {PAS_BLACK} !important;
        border-radius: 12px !important;
        font-weight: 900 !important;
    }}
    div.stButton > button[kind="secondary"], .stButton > button {{ min-height:52px !important; font-size:16px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}
    .stDownloadButton > button {{ min-height:62px !important; font-size:20px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}

    .kpi-card {{ background:#fff !important; border-radius:18px !important; border:1px solid #e4e7eb !important; box-shadow:0 5px 20px rgba(15,23,42,.08) !important; min-height:118px !important; padding:18px 22px !important; display:flex; align-items:center; gap:18px; }}
    .kpi-icon {{ width:64px; height:64px; border-radius:50%; background:#fff5bd; display:flex; align-items:center; justify-content:center; flex:none; }}
    .kpi-icon svg {{ width:35px; height:35px; stroke:#0A0A0A; stroke-width:2.5; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .kpi-label {{ color:#111 !important; font-size:15px !important; font-weight:950 !important; margin:0 0 3px !important; }}
    .kpi-value {{ color:#e9b900 !important; font-size:42px !important; line-height:.98 !important; font-weight:950 !important; text-shadow:none !important; }}
    .kpi-sub {{ color:#374151 !important; font-size:14px !important; margin-top:6px !important; }}

    .pas-results-title {{ color:#0A0A0A !important; font-size:28px !important; font-weight:950 !important; margin: 22px 0 8px !important; }}
    .pas-unmatched-pill {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:0 !important; border-radius:14px 14px 0 0 !important; padding:13px 20px !important; font-size:18px; font-weight:950; box-shadow:0 4px 14px rgba(0,0,0,.09); display:inline-block; }}
    .pas-table-wrap {{ background:#fff !important; border:1px solid #e0e4e9 !important; border-radius:0 16px 16px 16px !important; max-height:430px !important; overflow:auto !important; box-shadow:0 7px 25px rgba(15,23,42,.10) !important; }}
    table.pas-table {{ width:100%; border-collapse:collapse; font-size:14px !important; color:#0A0A0A !important; }}
    table.pas-table thead th {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:1px solid #e2ba00 !important; padding:12px 14px !important; font-weight:950 !important; position:sticky; top:0; z-index:5; text-align:left; white-space:nowrap; }}
    table.pas-table tbody td {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #e1e5eb !important; padding:10px 14px !important; vertical-align:top; }}
    table.pas-table tbody tr:nth-child(even) td {{ background:#fbfcfd !important; }}

    .pas-file-card {{ display:flex; align-items:center; gap:14px; background:#f4f6f8; border:1px solid #dfe4ea; border-radius:12px; padding:11px 14px; min-height:54px; margin: 4px 0 12px; }}
    .pas-file-icon {{ width:32px; height:32px; border-radius:8px; display:flex; align-items:center; justify-content:center; color:#fff; font-weight:950; font-size:11px; box-shadow:0 2px 8px rgba(0,0,0,.12); flex:none; }}
    .pas-file-icon.pdf {{ background:#c91d1d; }}
    .pas-file-main {{ flex:1; min-width:0; }}
    .pas-file-name {{ color:#0A0A0A; font-weight:950; font-size:15px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    .pas-file-size {{ color:#4b5563; font-weight:650; font-size:13px; margin-top:2px; }}
    .pas-file-check {{ width:24px; height:24px; border-radius:50%; background:#108a37; color:white; display:flex; align-items:center; justify-content:center; font-size:15px; font-weight:950; flex:none; }}
    .stCaption, div[data-testid="stCaptionContainer"], .stMarkdown p, .stInfo {{ color: #0A0A0A !important; }}
    div[data-testid="stAlert"], div[data-testid="stAlert"] * {{ color: #0A0A0A !important; }}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <style>
    .pas-sidebar-title {{ color:#fff; font-size:18px; font-weight:950; line-height:1.15; text-align:center; margin: 20px 0 8px; }}
    .pas-yellow-line {{ width:72px; height:4px; background:{PAS_YELLOW}; border-radius:99px; margin: 0 auto 22px; }}
    .pas-sidebar-copy {{ color:#fff !important; font-size:14px; line-height:1.52; font-weight:650; margin-bottom:24px; }}
    .pas-sidebar-rule {{ border-top:1px solid rgba(255,255,255,.22); margin:22px 0; }}
    .pas-sidebar-heading {{ color:{PAS_YELLOW}; font-size:19px; font-weight:950; margin: 0 0 16px; }}
    .pas-nav-row {{ display:grid; grid-template-columns: 26px 1fr; gap:10px; align-items:start; margin: 15px 0; color:#fff; font-weight:750; line-height:1.25; font-size:14px; }}
    .pas-nav-icon svg {{ width:21px; height:21px; stroke:{PAS_YELLOW}; stroke-width:2.4; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .pas-sidebar-footer {{ color:#fff; font-size:12px; font-weight:800; margin-top:28px; }}

    /* Bottom chase animation: small, low, runs once */
    .pas-bottom-chase-wrap {{ position: fixed; left: calc(18rem + 22px); right: 42px; bottom: 12px; height: 58px; pointer-events: none; z-index: 1; overflow: hidden; }}
    .pas-bottom-ground {{ position: absolute; left: 0; right: 0; bottom: 6px; border-bottom: 1px solid rgba(0,0,0,0.11); }}
    .pas-chase-pack {{ position: absolute; bottom: 8px; left: -150px; width: 150px; height: 48px; animation: pas-chase-run 13s linear 1 forwards; }}
    @keyframes pas-chase-run {{ 0% {{ transform: translateX(-120px); opacity: 0; }} 8% {{ opacity: 1; }} 88% {{ opacity: 1; }} 100% {{ transform: translateX(calc(100vw - 90px)); opacity: 0; }} }}
    .pas-truck-mini {{ position: absolute; left: 0; bottom: 5px; width: 54px; height: 30px; filter: drop-shadow(0 1px 1px rgba(0,0,0,.22)); }}
    .pas-truck-bed {{ position: absolute; left: 0; top: 5px; width: 34px; height: 19px; background: #FFD400; border: 3px solid #0A0A0A; border-radius: 4px 2px 3px 5px; transform: skewX(-10deg); }}
    .pas-truck-logo {{ position: absolute; left: 7px; top: 9px; font-size: 9px; font-weight: 950; color: #0A0A0A; line-height: 1; z-index: 3; }}
    .pas-truck-cab {{ position: absolute; left: 30px; top: 7px; width: 19px; height: 18px; background: #FFD400; border: 3px solid #0A0A0A; border-radius: 3px 5px 3px 2px; z-index: 2; }}
    .pas-truck-window {{ position: absolute; left: 34px; top: 10px; width: 7px; height: 7px; background: #a8d8e8; border: 2px solid #0A0A0A; border-radius: 2px; z-index: 4; }}
    .pas-truck-nose {{ position: absolute; left: 47px; top: 17px; width: 8px; height: 8px; background: #FFD400; border: 3px solid #0A0A0A; border-left: none; border-radius: 0 3px 3px 0; }}
    .pas-wheel {{ position: absolute; bottom: 0; width: 9px; height: 9px; background: #0A0A0A; border: 2px solid #222; border-radius: 50%; animation: pas-wheel-spin .32s linear infinite; z-index: 5; }}
    .pas-wheel::after {{ content: ""; position: absolute; inset: 2px; background: #FFD400; border-radius: 50%; }}
    .pas-wheel.back {{ left: 13px; }} .pas-wheel.front {{ left: 41px; }}
    @keyframes pas-wheel-spin {{ to {{ transform: rotate(360deg); }} }}
    .pas-speed-lines {{ position: absolute; left: -30px; top: 17px; width: 24px; height: 18px; }}
    .pas-speed-lines span {{ display:block; height:2px; background:#b9b9b9; margin:4px 0; border-radius:2px; animation: pas-flicker .55s linear infinite; }}
    .pas-speed-lines span:nth-child(2) {{ width: 16px; margin-left: 8px; }} .pas-speed-lines span:nth-child(3) {{ width: 11px; margin-left: 13px; }}
    @keyframes pas-flicker {{ 50% {{ opacity:.25; transform: translateX(-5px); }} }}
    .pas-dust {{ position:absolute; left:-5px; bottom:0; width:34px; height:14px; opacity:.75; }}
    .pas-dust span {{ position:absolute; bottom:0; background:#dac6a9; border-radius:50%; animation: pas-dust 1s linear infinite; }}
    .pas-dust span:nth-child(1) {{ width:12px; height:6px; left:0; }} .pas-dust span:nth-child(2) {{ width:16px; height:7px; left:10px; animation-delay:.2s; }} .pas-dust span:nth-child(3) {{ width:11px; height:5px; left:23px; animation-delay:.4s; }}
    @keyframes pas-dust {{ 50% {{ transform: translateX(-8px) scale(1.15); opacity:.4; }} }}
    .pas-stickman {{ position: absolute; left: 92px; bottom: 5px; width: 28px; height: 34px; animation: pas-runner-bob .35s ease-in-out infinite alternate; }}
    @keyframes pas-runner-bob {{ from {{ transform: translateY(1px); }} to {{ transform: translateY(-2px); }} }}
    .pas-stick-head {{ position:absolute; top:0; left:11px; width:8px; height:8px; border:2px solid #111; border-radius:50%; background:white; }}
    .pas-stick-body {{ position:absolute; left:15px; top:9px; width:2px; height:13px; background:#111; transform: rotate(12deg); transform-origin:top; }}
    .pas-stick-arm-a, .pas-stick-arm-b, .pas-stick-leg-a, .pas-stick-leg-b {{ position:absolute; width:2px; height:12px; background:#111; transform-origin:top; border-radius:2px; }}
    .pas-stick-arm-a {{ left:15px; top:11px; transform: rotate(58deg); animation: pas-arm-a .35s linear infinite alternate; }}
    .pas-stick-arm-b {{ left:15px; top:11px; transform: rotate(-50deg); animation: pas-arm-b .35s linear infinite alternate; }}
    .pas-stick-leg-a {{ left:16px; top:21px; height:14px; transform: rotate(48deg); animation: pas-leg-a .35s linear infinite alternate; }}
    .pas-stick-leg-b {{ left:16px; top:21px; height:14px; transform: rotate(-42deg); animation: pas-leg-b .35s linear infinite alternate; }}
    @keyframes pas-arm-a {{ to {{ transform: rotate(-45deg); }} }} @keyframes pas-arm-b {{ to {{ transform: rotate(55deg); }} }} @keyframes pas-leg-a {{ to {{ transform: rotate(-45deg); }} }} @keyframes pas-leg-b {{ to {{ transform: rotate(48deg); }} }}
    </style>
    """,
    unsafe_allow_html=True,
)


def render_bottom_chase():
    st.markdown(
        """
        <div class="pas-bottom-chase-wrap" aria-hidden="true">
            <div class="pas-bottom-ground"></div>
            <div class="pas-chase-pack">
                <div class="pas-speed-lines"><span></span><span></span><span></span></div>
                <div class="pas-dust"><span></span><span></span><span></span></div>
                <div class="pas-truck-mini">
                    <div class="pas-truck-bed"></div><div class="pas-truck-logo">PAS</div><div class="pas-truck-cab"></div><div class="pas-truck-window"></div><div class="pas-truck-nose"></div><div class="pas-wheel back"></div><div class="pas-wheel front"></div>
                </div>
                <div class="pas-stickman"><div class="pas-stick-head"></div><div class="pas-stick-body"></div><div class="pas-stick-arm-a"></div><div class="pas-stick-arm-b"></div><div class="pas-stick-leg-a"></div><div class="pas-stick-leg-b"></div></div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_selected_file_card(uploaded_file):
    size = getattr(uploaded_file, "size", 0) or 0
    size_text = f"{size / (1024 * 1024):.1f} MB" if size >= 1024 * 1024 else f"{size / 1024:.0f} KB"
    st.markdown(
        f'''
        <div class="pas-file-card">
            <div class="pas-file-icon pdf">PDF</div>
            <div class="pas-file-main"><div class="pas-file-name">{escape(getattr(uploaded_file, "name", "Uploaded file"))}</div><div class="pas-file-size">{size_text}</div></div>
            <div class="pas-file-check">✓</div>
        </div>
        ''',
        unsafe_allow_html=True,
    )


with st.sidebar:
    if Path("pas_logo.png").exists():
        st.image("pas_logo.png", use_column_width=True)
    elif Path("assets/pas_logo.png").exists():
        st.image("assets/pas_logo.png", use_column_width=True)
    else:
        st.markdown('<div style="background:#FFD400;color:#000;border-radius:14px;padding:18px;text-align:center;font-weight:950;font-size:30px;">PAS</div>', unsafe_allow_html=True)
    st.markdown(
        """
        <div class="pas-sidebar-title">PAS Vendor<br>PDF Converter</div>
        <div class="pas-yellow-line"></div>
        <div class="pas-sidebar-copy">Upload vendor hire report PDFs, then export one clean standard Excel file.</div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-heading">Instructions</div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M16 16l-4-4-4 4"/><path d="M12 12v9"/><path d="M20 16.6A5 5 0 0 0 18 7h-1.3A8 8 0 1 0 4 15.3"/></svg></span><span>Upload Vendor PDF Reports</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M5 3l14 9-14 9V3z"/></svg></span><span>Run PDF Converter</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><path d="M7 10l5 5 5-5"/><path d="M12 15V3"/></svg></span><span>Download Clean Excel</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M9 3h6"/><path d="M10 3v5l-4 8a4 4 0 0 0 3.6 5.7h4.8A4 4 0 0 0 18 16l-4-8V3"/><path d="M8 14h8"/></svg></span><span>Standard PAS Columns</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="M21 21l-4.3-4.3"/></svg></span><span>Smoke Crack</span></div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-footer">PAS NW Ltd • v1.0 Prototype Build</div>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    f"""
    <div class="pas-hero">
      <div class="pas-hero-logo">PAS</div>
      <div class="pas-hero-text">PAS NW Ltd<span class="pas-hero-dot">•</span><span class="pas-hero-version">{APP_VERSION}</span></div>
    </div>
    """,
    unsafe_allow_html=True,
)

# =========================
# Extraction helpers
# =========================
def clean_cell(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).replace("\u00a0", " ").strip()
    return re.sub(r"\s+", " ", text)


def clean_money(value) -> str:
    text = clean_cell(value).replace(",", "")
    if not text:
        return ""
    m = re.search(r"£?\s*(-?\d+(?:\.\d{1,2})?)", text)
    if not m:
        return ""
    return f"{float(m.group(1)):.2f}"


def clean_qty(value) -> str:
    text = clean_cell(value)
    if not text:
        return ""
    try:
        f = float(text)
        return str(int(f)) if f.is_integer() else str(f)
    except Exception:
        return text


def normalise_date(value, report_year: Optional[int] = None) -> str:
    text = clean_cell(value)
    if not text:
        return ""
    text = text.replace("00:00:00", "").strip()
    if re.match(r"^\d{1,2}-[A-Za-z]{3}-$", text) and report_year:
        text = f"{text}{report_year}"
    if re.match(r"^\d{1,2}-[A-Za-z]{3}$", text) and report_year:
        text = f"{text}-{report_year}"
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return text
    return parsed.strftime("%d/%m/%Y")


def extract_report_year(text: str) -> Optional[int]:
    m = re.search(r"(?:DATE:|on|Run by|Page)\s*[^\n\r]*(\d{4})", text, re.I)
    if m:
        return int(m.group(1))
    m = re.search(r"\b(20\d{2})\b", text)
    return int(m.group(1)) if m else None


def extract_pdf_text(uploaded_file) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        from PyPDF2 import PdfReader
    pdf_bytes = uploaded_file.read()
    uploaded_file.seek(0)
    reader = PdfReader(io.BytesIO(pdf_bytes))
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        pages.append(f"\n--- PAGE {i} ---\n{page_text}")
    return "\n".join(pages)


def make_row(site="", order_no="", contract_no="", fleet="", desc="", qty="", on_hire="", weekly="", net="") -> Dict[str, str]:
    weekly_clean = clean_money(weekly)
    net_clean = clean_money(net) or weekly_clean
    return {
        "Site": clean_cell(site),
        "Order No": clean_cell(order_no).replace(" / ", "/").replace("/ ", "/").replace(" /", "/"),
        "Contract No": clean_cell(contract_no),
        "Fleet / Stock No": clean_cell(fleet),
        "Description": clean_cell(desc),
        "Qty": clean_qty(qty),
        "On Hire Date": normalise_date(on_hire),
        "Weekly Rate": weekly_clean,
        "Net Weekly": net_clean,
    }


def valid_row(row: Dict[str, str]) -> bool:
    return bool(row.get("Order No") and row.get("Description") and row.get("Qty") and row.get("On Hire Date"))


def dedupe_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    seen = set()
    out = []
    for r in rows:
        key = tuple(clean_cell(r.get(c, "")).lower() for c in OUTPUT_COLUMNS)
        if key in seen:
            continue
        seen.add(key)
        out.append({c: r.get(c, "") for c in OUTPUT_COLUMNS})
    return out


# =========================
# Smart extraction patterns
# These are not fixed page templates. They use order/date/contract/rate anchors
# and preserve parent contract/site context for wrapped PDF text.
# =========================
def parse_ambrose_live_hire(text: str) -> List[Dict[str, str]]:
    rows = []
    report_year = extract_report_year(text)
    current_contract = ""
    current_order = ""
    current_site = ""
    address_parts = []

    # PDF text extraction can flip this heading in either direction:
    #   Order No: P141/H5660 Delivery Address ...
    # or
    #   P141/H5660Order No: Delivery Address ...
    # so the order number is captured from either side of the Order No label.
    contract_re = re.compile(
        r"^(?P<contract>\d{6,12})\s+.*?"
        r"(?:Order\s+No:\s*(?P<order_after>P[A-Za-z0-9&]*(?:\s*/\s*H?\d+)?)\s+Delivery\s+Address|"
        r"(?P<order_before>P[A-Za-z0-9&]*(?:\s*/\s*H?\d+)?)\s*Order\s+No:\s+Delivery\s+Address)"
        r"\s*(?P<addr>.*)$",
        re.I,
    )
    item_re = re.compile(
        r"^(?P<fleet>[A-Z0-9][A-Z0-9/-]{1,24})\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<qty>-?\d+(?:\.\d+)?)\s+"
        r"(?P<weekly>-?\d+(?:\.\d+)?)\s+"
        r"(?P<discount>-?\d+(?:\.\d+)?%)\s+"
        r"(?P<net>-?\d+(?:\.\d+)?)\s+"
        r"(?P<lastinv>N/A|\d{1,2}/\d{1,2}/\d{2,4})\s+"
        r"(?P<onhire>\d{1,2}/\d{1,2}/\d{2,4})(?:\s+\d{1,2}:\d{2}:\d{2})?\s+"
        r"(?P<type>[A-Za-z]+)\s*$",
        re.I,
    )

    def commit_site():
        nonlocal current_site, address_parts
        if address_parts:
            current_site = clean_cell(" ".join(address_parts)).strip(" ,")
            address_parts = []

    for raw in text.splitlines():
        line = clean_cell(raw)
        if not line:
            continue
        low = line.lower()
        if (
            "live hire detail report" in low
            or low.startswith("item no description")
            or low.startswith("contract no acct no")
            or low.startswith("depot:")
            or re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}\s+page\s+\d+", low)
            or line.startswith("--- PAGE")
        ):
            continue

        cm = contract_re.match(line)
        if cm:
            commit_site()
            current_contract = cm.group("contract")
            current_order = clean_cell(cm.group("order_after") or cm.group("order_before") or "")
            current_site = ""
            address_parts = [cm.group("addr")] if cm.group("addr") else []
            continue

        im = item_re.match(line)
        if im and current_order:
            commit_site()
            rows.append(make_row(
                site=current_site,
                order_no=current_order,
                contract_no=current_contract,
                fleet=im.group("fleet"),
                desc=im.group("desc"),
                qty=im.group("qty"),
                on_hire=normalise_date(im.group("onhire"), report_year),
                weekly=im.group("weekly"),
                net=im.group("net"),
            ))
            continue

        if current_order and not re.match(r"^\d{6,12}\s+", line) and not item_re.match(line):
            if not re.search(r"\b(Page|Customer Range|PAS002|Type)\b", line, re.I):
                address_parts.append(line)

    return [r for r in rows if valid_row(r)]


def parse_mcs_items_on_hire(text: str) -> List[Dict[str, str]]:
    rows = []
    report_year = extract_report_year(text) or datetime.now().year
    current_contract = ""
    current_order = ""
    current_site = ""
    contract_re = re.compile(r"^Contract\s+No:\s*(?P<contract>\S+)\s+PO\s+Number:\s*(?P<order>.+)$", re.I)
    site_re = re.compile(r"^Site\s+Name:\s*(?P<site>.+)$", re.I)
    item_re = re.compile(
        r"^(?P<date>\d{1,2}-[A-Za-z]{3}(?:-\d{2,4})?)\s+"
        r"(?P<fleet>[A-Za-z0-9/-]+)\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<qty>-?\d+(?:\.\d+)?)\s*$"
    )
    for raw in text.splitlines():
        line = clean_cell(raw)
        if not line:
            continue
        cm = contract_re.match(line)
        if cm:
            current_contract = cm.group("contract")
            current_order = cm.group("order")
            current_site = ""
            continue
        sm = site_re.match(line)
        if sm:
            current_site = sm.group("site")
            continue
        if line.lower().startswith(("date:", "cust job", "items on hire", "customer no", "address:", "page ")):
            continue
        im = item_re.match(line)
        if im and current_order:
            rows.append(make_row(
                site=current_site,
                order_no=current_order,
                contract_no=current_contract,
                fleet=im.group("fleet"),
                desc=im.group("desc"),
                qty=im.group("qty"),
                on_hire=normalise_date(im.group("date"), report_year),
                weekly="",
                net="",
            ))
    return [r for r in rows if valid_row(r)]


def parse_equipment_custom(text: str) -> List[Dict[str, str]]:
    rows = []
    report_year = extract_report_year(text)
    current_site = ""
    awaiting_site = False
    # Supports: CAT STOCK SERIAL QTY DESCRIPTION DATE CONTRACT ORDER RATE/w
    line_re = re.compile(
        r"^(?P<cat>[A-Z0-9][A-Z0-9/-]{1,20})\s+"
        r"(?P<stock>[A-Z0-9][A-Z0-9/-]{1,30})\s+"
        r"(?P<serial>[A-Z0-9][A-Z0-9/-]{1,35})\s+"
        r"(?P<qty>-?\d+(?:\.\d+)?)\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<date>\d{1,2}/\d{1,2}/\d{2,4}|\d{1,2}-[A-Za-z]{3}-\d{2,4})\s*(?:\d{1,2}:\d{2}:\d{2})?\s+"
        r"(?P<contract>[A-Za-z0-9-]+)\s+"
        r"(?P<order>[A-Za-z0-9/& .-]+?)\s+"
        r"(?P<rate>£?\d+(?:\.\d{1,2})?)\s*/?\s*w?\s*$",
        re.I,
    )
    for raw in text.splitlines():
        line = clean_cell(raw)
        if not line:
            continue
        low = line.lower()
        if low.startswith("site:"):
            site_after = clean_cell(line.split(":", 1)[1]) if ":" in line else ""
            if site_after:
                current_site = site_after
                awaiting_site = False
            else:
                awaiting_site = True
            continue
        if awaiting_site:
            if not re.search(r"cat\s+stock|pas \(nw\)|serial no|hire rate", low, re.I):
                current_site = line
                awaiting_site = False
                continue
        if re.search(r"cat\s+stock|serial no|hire rate|equipment currently|page \d+", low, re.I):
            continue
        m = line_re.match(line)
        if m:
            rows.append(make_row(
                site=current_site,
                order_no=m.group("order"),
                contract_no=m.group("contract"),
                fleet=m.group("stock"),
                desc=m.group("desc"),
                qty=m.group("qty"),
                on_hire=normalise_date(m.group("date"), report_year),
                weekly=m.group("rate"),
                net=m.group("rate"),
            ))
    return [r for r in rows if valid_row(r)]


def choose_stock_and_description(prefix: str) -> Tuple[str, str]:
    """Best-effort split for reports where stock/serial columns collapse in PDF text."""
    tokens = prefix.split()
    if len(tokens) < 3:
        return "", prefix
    category = tokens[0]
    rest = tokens[1:]

    # Description usually starts at a recognisable plant/equipment word or at first mostly textual phrase.
    desc_start_words = {
        "pallet", "forks", "pipe", "pusher", "spike", "drive", "unit", "rock", "auger", "pile", "cutter",
        "13-20t", "5-10t", "5-8t", "75mm", "pd7", "towable", "solar", "event", "toilet",
    }
    desc_idx = None
    for i, tok in enumerate(rest):
        t = tok.lower().strip("()")
        if t in desc_start_words or re.search(r"mm$|cm$|tonne|fork|pallet|pipe|drive|cutter", t):
            desc_idx = i
            break
    if desc_idx is None:
        # Fallback: stock is generally one or two tokens after category.
        desc_idx = 2 if len(rest) > 3 else 1
    stock = " ".join(rest[:desc_idx]).strip()
    desc = " ".join(rest[desc_idx:]).strip()
    if not stock:
        stock = category
    return stock, desc


def parse_current_hires_anchored(text: str) -> List[Dict[str, str]]:
    rows = []
    report_year = extract_report_year(text)
    current_site = ""
    date_re = re.compile(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b")
    # Prefix [qty] [date] [contract] [order] [rate] [period]
    row_re = re.compile(
        r"^(?P<prefix>.+?)\s+"
        r"(?P<qty>-?\d+(?:\.\d+)?)\s+"
        r"(?P<date>\d{1,2}/\d{1,2}/\d{2,4})\s+"
        r"(?P<contract>[A-Za-z]{1,4}\d{3,})\s+"
        r"(?P<order>[A-Za-z0-9/& .-]+?)\s+"
        r"£?\s*(?P<rate>\d+(?:\.\d{1,2})?)\s+"
        r"(?P<period>\S+)\s*$",
        re.I,
    )
    for raw in text.splitlines():
        line = clean_cell(raw)
        if not line:
            continue
        low = line.lower()
        if line.startswith("--- PAGE") or low.startswith(("current hires", "page ", "pas (nw)", "category stock", "rockbreakers", "criteria:")):
            continue
        if re.search(r"category\s+stock\s+no\s+serial\s+no\s+description", low):
            continue
        m = row_re.match(line)
        if m:
            stock, desc = choose_stock_and_description(m.group("prefix"))
            rows.append(make_row(
                site=current_site,
                order_no=m.group("order"),
                contract_no=m.group("contract"),
                fleet=stock,
                desc=desc,
                qty=m.group("qty"),
                on_hire=normalise_date(m.group("date"), report_year),
                weekly=m.group("rate"),
                net=m.group("rate"),
            ))
            continue
        # Site headings are lines between table sections, usually uppercase/mixed case with no date/rate/order.
        if not date_re.search(line) and not re.search(r"£|\bWHC\d+\b|\bHC\d+\b|\bP\d{2,4}/H\d+", line, re.I):
            if len(line) > 3 and not re.search(r"run by|customer|address|name:|date:", low):
                current_site = line
    return [r for r in rows if valid_row(r)]


def parse_generic_order_blocks(text: str) -> List[Dict[str, str]]:
    """Fallback smart parser: preserve Contract/Order context and read rows by date/order anchors."""
    rows = []
    report_year = extract_report_year(text) or datetime.now().year
    current_site = ""
    current_contract = ""
    current_order = ""
    contract_patterns = [
        re.compile(r"Contract\s+No:?\s*(?P<contract>[A-Za-z0-9-]+).*?(?:PO|Order)\s+(?:Number|No):?\s*(?P<order>[A-Za-z0-9/& .-]+)", re.I),
        re.compile(r"(?P<contract>\d{6,12}).*?Order\s+No:?\s*(?P<order>[A-Za-z0-9/& .-]+)", re.I),
    ]
    site_patterns = [
        re.compile(r"Site\s+Name:?\s*(?P<site>.+)", re.I),
        re.compile(r"Site:?\s*(?P<site>.+)", re.I),
        re.compile(r"Delivery\s+Address\s*(?P<site>.+)", re.I),
    ]
    date_row_re = re.compile(r"^(?P<date>\d{1,2}[-/]\w{3,9}[-/]?\d{0,4}|\d{1,2}/\d{1,2}/\d{2,4})\s+(?P<body>.+?)\s+(?P<qty>-?\d+(?:\.\d+)?)$")
    for raw in text.splitlines():
        line = clean_cell(raw)
        if not line:
            continue
        for cp in contract_patterns:
            cm = cp.search(line)
            if cm:
                current_contract = cm.group("contract")
                current_order = cm.group("order")
                break
        for sp in site_patterns:
            sm = sp.search(line)
            if sm:
                val = clean_cell(sm.group("site"))
                if val and not val.lower().startswith("name:"):
                    current_site = val
                break
        rm = date_row_re.match(line)
        if rm and current_order:
            body = rm.group("body")
            bits = body.split()
            fleet = bits[0] if bits else ""
            desc = " ".join(bits[1:]) if len(bits) > 1 else body
            rows.append(make_row(
                site=current_site,
                order_no=current_order,
                contract_no=current_contract,
                fleet=fleet,
                desc=desc,
                qty=rm.group("qty"),
                on_hire=normalise_date(rm.group("date"), report_year),
                weekly="",
                net="",
            ))
    return [r for r in rows if valid_row(r)]


def smart_extract_vendor_pdf(uploaded_file) -> Tuple[pd.DataFrame, List[str]]:
    text = extract_pdf_text(uploaded_file)
    warnings = []
    all_rows: List[Dict[str, str]] = []

    parsers = [
        ("Contract block hire report", parse_ambrose_live_hire),
        ("Items on hire contract blocks", parse_mcs_items_on_hire),
        ("Equipment on hire table", parse_equipment_custom),
        ("Current hires anchored table", parse_current_hires_anchored),
        ("Generic order/date fallback", parse_generic_order_blocks),
    ]

    parser_counts = {}
    for name, parser in parsers:
        try:
            found = parser(text)
            found = [r for r in found if valid_row(r)]
            parser_counts[name] = len(found)
            all_rows.extend(found)
        except Exception as exc:
            warnings.append(f"{name} parser skipped: {exc}")
            parser_counts[name] = 0

    rows = dedupe_rows(all_rows)
    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    if df.empty:
        warnings.append("No rows were extracted. This PDF may be scanned/image-only or needs another sample-specific rule.")
    else:
        best = max(parser_counts.items(), key=lambda x: x[1]) if parser_counts else ("", 0)
        if best[1] > 0:
            warnings.append(f"Best extraction route: {best[0]} ({best[1]} rows before de-duplication).")
    return df, warnings


def make_excel(df: pd.DataFrame, summary_df: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        df.to_excel(writer, index=False, sheet_name="Converted Report")
        wb = writer.book
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="FFD400")
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 25
            ws.auto_filter.ref = ws.dimensions
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(name="Calibri", size=10, bold=(cell.row == 1), color="000000")
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = border
                    if cell.row == 1:
                        cell.fill = header_fill
            for column_cells in ws.columns:
                letter = column_cells[0].column_letter
                max_len = 0
                for cell in column_cells:
                    max_len = max(max_len, len(clean_cell(cell.value)))
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)
            # date/rate widths
            for col in range(1, ws.max_column + 1):
                header = clean_cell(ws.cell(1, col).value)
                if header in {"Description", "Site"}:
                    ws.column_dimensions[ws.cell(1, col).column_letter].width = 38
                elif header in {"Weekly Rate", "Net Weekly"}:
                    ws.column_dimensions[ws.cell(1, col).column_letter].width = 14
    return out.getvalue()


def render_table(df: pd.DataFrame):
    if df is None or df.empty:
        st.markdown('<div class="pas-unmatched-pill">Converted Rows</div>', unsafe_allow_html=True)
        st.markdown('<div class="pas-table-wrap"><table class="pas-table"><tbody><tr><td>No rows found yet.</td></tr></tbody></table></div>', unsafe_allow_html=True)
        return
    display_df = df[OUTPUT_COLUMNS].head(300).copy()
    header_html = "".join(f"<th>{escape(c)}</th>" for c in display_df.columns)
    rows_html = []
    for _, row in display_df.iterrows():
        rows_html.append("<tr>" + "".join(f"<td>{escape(clean_cell(row.get(c, '')))}</td>" for c in display_df.columns) + "</tr>")
    st.markdown('<div class="pas-unmatched-pill">Converted Rows</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="pas-table-wrap"><table class="pas-table"><thead><tr>{header_html}</tr></thead><tbody>{"".join(rows_html)}</tbody></table></div>', unsafe_allow_html=True)

# =========================
# Main UI
# =========================
st.markdown('<div class="pas-upload-card"><div class="pas-upload-title">Upload Vendor Hire Report PDF(s)</div>', unsafe_allow_html=True)
pdf_files = st.file_uploader("Upload Vendor Hire Report PDF(s)", type=["pdf"], accept_multiple_files=True, label_visibility="collapsed", key="pdf_upload")
if pdf_files:
    for f in pdf_files:
        render_selected_file_card(f)
st.markdown('</div>', unsafe_allow_html=True)

run = st.button("▶  Run PDF converter", use_container_width=True)

if "vendor_pdf_converter_results" not in st.session_state:
    st.session_state["vendor_pdf_converter_results"] = None

if run:
    if not pdf_files:
        st.warning("Please upload at least one vendor hire report PDF.")
        st.stop()
    all_frames = []
    log_rows = []
    with st.spinner("Reading PDFs and converting hire lines..."):
        for uploaded in pdf_files:
            try:
                df, warnings = smart_extract_vendor_pdf(uploaded)
                if not df.empty:
                    df.insert(0, "Source File", uploaded.name)
                    all_frames.append(df)
                log_rows.append({
                    "File": uploaded.name,
                    "Rows Extracted": len(df),
                    "Notes": " | ".join(warnings),
                })
            except Exception as exc:
                log_rows.append({"File": uploaded.name, "Rows Extracted": 0, "Notes": f"Failed: {exc}"})

    if all_frames:
        combined = pd.concat(all_frames, ignore_index=True)
        # Output sheet must use required columns only. Source file is kept out of final output columns.
        final_df = combined[OUTPUT_COLUMNS].copy()
    else:
        final_df = pd.DataFrame(columns=OUTPUT_COLUMNS)

    summary_df = pd.DataFrame(log_rows)
    excel_bytes = make_excel(final_df, summary_df)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    st.session_state["vendor_pdf_converter_results"] = {
        "df": final_df,
        "summary": summary_df,
        "excel_bytes": excel_bytes,
        "filename": f"PAS_Vendor_PDF_Converted_{stamp}.xlsx",
        "total_files": len(pdf_files),
        "total_rows": len(final_df),
        "with_fleet": int(final_df["Fleet / Stock No"].replace("", pd.NA).dropna().shape[0]) if not final_df.empty else 0,
        "with_rate": int(final_df["Weekly Rate"].replace("", pd.NA).dropna().shape[0]) if not final_df.empty else 0,
    }

results = st.session_state.get("vendor_pdf_converter_results")

if results is not None:
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><path d="M14 2v6h6"/><path d="M9 13h6"/><path d="M9 17h6"/></svg></div><div><div class="kpi-label">PDFs processed</div><div class="kpi-value">{results["total_files"]}</div><div class="kpi-sub">Vendor reports</div></div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M3 6h18"/><path d="M3 12h18"/><path d="M3 18h18"/></svg></div><div><div class="kpi-label">Rows converted</div><div class="kpi-value">{results["total_rows"]}</div><div class="kpi-sub">Clean Excel lines</div></div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M4 7h16"/><path d="M4 12h16"/><path d="M4 17h10"/><circle cx="18" cy="17" r="2"/></svg></div><div><div class="kpi-label">Fleet numbers</div><div class="kpi-value">{results["with_fleet"]}</div><div class="kpi-sub">Captured where supplied</div></div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M12 1v22"/><path d="M17 5H9.5a3.5 3.5 0 0 0 0 7H14a3.5 3.5 0 0 1 0 7H6"/></svg></div><div><div class="kpi-label">Rates captured</div><div class="kpi-value">{results["with_rate"]}</div><div class="kpi-sub">Blank if not shown</div></div></div>', unsafe_allow_html=True)

    st.markdown('<div class="pas-results-title">Results</div>', unsafe_allow_html=True)
    render_table(results["df"])

    with st.expander("Extraction notes"):
        st.dataframe(results["summary"], use_container_width=True, hide_index=True)

    dl_left, dl_mid, dl_right = st.columns([1.3, 1, 1.3])
    with dl_mid:
        st.download_button(
            "⬇  Download Excel",
            data=results["excel_bytes"],
            file_name=results["filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
else:
    st.info("Upload vendor hire report PDFs, then click Run PDF converter.")

if "animation_shown" not in st.session_state:
    render_bottom_chase()
    st.session_state["animation_shown"] = True
